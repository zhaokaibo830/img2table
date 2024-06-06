import json
import math
from typing import Dict, List, Tuple, Any

from openpyxl import load_workbook

def excel_to_json(excel_path: str) -> Dict:
    """
    输入一个excel文件，输出对应的json统一格式描述，excel是xlsx格式
    :return:
    """
    # 读取 Excel 文件
    wb = load_workbook(excel_path)
    sheet = wb.worksheets[0]

    # 转换为 JSON 格式
    cells = []

    # 遍历每个单元格
    for row in sheet.iter_rows():
        for cell in row:
            # 获取单元格的值
            cell_value = cell.value

            # 检查单元格是否为空
            if cell_value is None:
                continue

            # 初始化单元格的跨度
            col_start = cell.column
            col_span = [col_start, col_start]
            row_span = [cell.row, cell.row]

            # 判断单元格是否为合并单元格
            for merged_cell_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_cell_range:
                    # 获取合并单元格的起始行号和列号
                    merged_start_cell = merged_cell_range.min_row, merged_cell_range.min_col
                    merged_end_cell = merged_cell_range.max_row, merged_cell_range.max_col
                    row_span = [merged_start_cell[0], merged_end_cell[0]]
                    col_span = [merged_start_cell[1], merged_end_cell[1]]
                    break

            # 检查单元格的颜色是否为特定的 RGB（200，200，200）
            cell_color = cell.fill.start_color.rgb
            if cell_color == "FFC8C8C8":
                node_type = "key"
            else:
                node_type = "value"

            # 添加单元格到列表中
            cells.append({
                "colspan": col_span,
                "rowspan": row_span,
                "text": str(cell_value).strip().replace("\n", " "),
                "node_type": node_type
            })


    return {"content": cells}



if __name__ == '__main__':
    excel_path = r"E:\code\image2table\output\2\[0, 0, 919, 217]_0.xlsx"
    output_json_path = "output.json"
    table=excel_to_json(excel_path)
    print(table)
    # word_to_json(docx_file)
    # image_to_json(image_path)
    # any_format_to_json(docx_file)
